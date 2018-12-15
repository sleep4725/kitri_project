from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import time
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
import os
import json
from collections import OrderedDict
import sys
# ______________________________________________________
class STU:
    def __init__(self):
        # 작업 디렉토리 이동 ____________________
        try:
            os.chdir("C:\\Users\\sleep\\Desktop\\Today_01")
        except:
            os.mkdir("C:\\Users\\sleep\\Desktop\\Today_01")
            os.chdir("C:\\Users\\sleep\\Desktop\\Today_01")
        finally:
            print ("current directory : {}".format(os.curdir))
        self.target_url = "https://movie.naver.com/"
        #-------------------------------------
        self.options = Options()# 객체 생성
        self.options.add_argument("--headless")
        # self.options.binary_location = "C:\\Users\\sleep\\Desktop\\chrom_driver\\chromedriver.exe"
        # -------------------------------------
        self.driver = webdriver.Chrome \
            (executable_path="C:\\Users\\sleep\\Desktop\\chrom_driver\\chromedriver.exe",
             chrome_options=self.options)

        self.currentMov = {
            "reservMovi":{},   # 현재 상영영화 - 예매순 : reservMovi
            "releaseMovi":{},  # 현재 상영영화 - 개봉순 : releaseMovi
            "gradeMovi":{},    # 현재 상영영화 - 평점순 : gradeMovi
            "likeMovi":{},     # 현재 상영영화 - 좋아요순 : gradeMovi
        }
    # Instance method (1)
    def urlRequests(self):
        self.driver.get(self.target_url)
        assert "네이버 영화" in self.driver.title
        print (self.driver.title)
        time.sleep(2)
        # 상영작 예정작
        # menu02_on
        self.driver.find_element_by_xpath(
            '//*[@id="scrollbar"]/div[1]/div/div/ul/li[2]/a/strong').click()
        assert "현재 상영영화 : 네이버 영화" in self.driver.title
        print (self.driver.title)

    def doCommon(self, mark):
        bsObject = BeautifulSoup(self.driver.page_source, "html.parser")
        mvList = bsObject.select('dt.tit > a')
        for n, i in enumerate(mvList):
            self.currentMov[mark][n + 1] = i.string
            # print(i.string)

    def doExcel(self, mark, content):
        # excel data _________________________
        try:
            self.workBook = load_workbook('C:\\Users\\sleep\\Desktop\\Today_01\\mv.xlsx')
        except FileNotFoundError as e:
            try:
                self.workBook = Workbook()
                self.workBook.save(filename='C:\\Users\\sleep\\Desktop\\Today_01\\mv.xlsx')
            except:
                print ("excel file create error ")
                sys.exit(1)
            else:
                self.workBook = load_workbook('C:\\Users\\sleep\\Desktop\\Today_01\\mv.xlsx')

        # excel write_
        NUM_INDX = ['B',3]
        MOV_CONT = ['C',3]
        # Sheet 생성___
        wsheet = self.workBook.create_sheet(title=content)
        # Cell 열너비 조정
        wsheet.column_dimensions['C'].width = 47.4
        # Cell 데이터 적재
        wsheet[NUM_INDX[0] + str(NUM_INDX[1])] = "순위"
        wsheet[MOV_CONT[0] + str(MOV_CONT[1])] = "영화이름"
        # 내용입력
        for i in self.currentMov["reservMovi"].keys():
            # 인덱스 증가____________________________
            NUM_INDX[1] += 1
            MOV_CONT[1] += 1
            # _____________________________________
            wsheet[NUM_INDX[0] + str(NUM_INDX[1])] = i  # 순위
            wsheet[MOV_CONT[0] + str(MOV_CONT[1])] = self.currentMov[mark][i]  # 내용
        self.workBook.save("C:\\Users\\sleep\\Desktop\\Today_01\\mv.xlsx")

    # Instance method (2)
    def requestReserve(self):
        # 예매순
        # //*[@id="content"]/div[1]/div[1]/div[2]/ul[2]/li[1]/a
        self.driver.find_element_by_xpath(
            '//*[@id="content"]/div[1]/div[1]/div[2]/ul[2]/li[1]/a').click();time.sleep(2)
        print ("예매순 진행 중 ... ")
        self.doCommon("reservMovi")
        self.doExcel('reservMovi', '예매순')

    # Instance method (3)
    def requestRelease(self):
        # 개봉순
        # //*[@id="content"]/div[1]/div[1]/div[2]/ul[2]/li[2]/a
        self.driver.find_element_by_xpath(
            '//*[@id="content"]/div[1]/div[1]/div[2]/ul[2]/li[2]/a').click();time.sleep(2)
        # #content > div.article > div:nth-child(1) > div.lst_wrap > ul > li:nth-child(1) > dl > dt > a
        print("개봉순 진행 중 ... ")
        self.doCommon('releaseMovi')
        self.doExcel('releaseMovi', '개봉순')

    # Instance method (4)
    def requestGradeMovi(self):
        # 평점순
        # //*[@id="content"]/div[1]/div[1]/div[2]/ul[2]/li[3]/a
        self.driver.find_element_by_xpath(
            '//*[@id="content"]/div[1]/div[1]/div[2]/ul[2]/li[3]/a').click();time.sleep(2)
        print("평점순 진행 중 ... ")
        self.doCommon('gradeMovi')
        self.doExcel('gradeMovi', '평점순')

    # Instance method (5)
    def requestlikeMovi(self):
        # 좋아요 순
        # //*[@id="content"]/div[1]/div[1]/div[2]/ul[2]/li[4]/a
        self.driver.find_element_by_xpath(
            '//*[@id="content"]/div[1]/div[1]/div[2]/ul[2]/li[4]/a').click();time.sleep(2)
        print("좋아요 순 진행 중 ... ")
        self.doCommon('likeMovi')
        self.doExcel('likeMovi', '좋아요 순')

    # Instance method (6)
    def jsonFileCreate(self):
        fileData = OrderedDict()
        fileData['Reserve']   = self.currentMov['reservMovi']    # 예매순
        fileData['Release']   = self.currentMov['releaseMovi']   # 개봉순
        fileData['GradeMovi'] = self.currentMov['gradeMovi']     # 평점순
        fileData['likeMovi']  = self.currentMov['likeMovi']      # 좋아요순
        with open('movie.json', 'w', encoding='utf-8') as makeJson:
            json.dump(fileData, makeJson, ensure_ascii=False, indent="\t")

def main():
    sNode = STU()
    sNode.urlRequests()
    sNode.requestReserve()      # 예매순
    sNode.requestRelease()      # 개봉순
    sNode.requestGradeMovi()    # 평점순
    sNode.requestlikeMovi()     # 좋아요 순
    sNode.jsonFileCreate()      # json
if __name__ == "__main__":
    main()